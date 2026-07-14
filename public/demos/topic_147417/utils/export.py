"""
导出工具模块
支持 Excel（双 Sheet 美化）、CSV 导出
字段：商品名称 / 品牌 / 型号 / 数量 / 单价 / 小计 / 类型 / 供电方式 / 通信协议 /
      零火线要求 / 安装方式 / 购买链接 / 备注
"""
import csv
import io
import os
import sys
from datetime import datetime
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 确保项目根目录在 Python 路径中（用于 import helpers）
_project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)

from utils.helpers import format_device_for_export, detect_conflicts

# ============================================
# 导出列定义（顺序固定）
# ============================================
EXPORT_COLUMNS = [
    '商品名称', '品牌', '型号', '数量', '单价(元)', '小计(元)',
    '类型', '供电方式', '通信协议', '零火线要求', '安装方式',
    '购买链接', '备注',
]

# 列宽配置
COLUMN_WIDTHS = {
    '商品名称': 25, '品牌': 15, '型号': 20, '数量': 8,
    '单价(元)': 12, '小计(元)': 12, '类型': 12, '供电方式': 12,
    '通信协议': 15, '零火线要求': 14, '安装方式': 12,
    '购买链接': 40, '备注': 20,
}


def _prepare_export_rows(devices):
    """将设备列表转换为导出行列表（使用 format_device_for_export）。"""
    rows = []
    for device in devices:
        if not isinstance(device, dict):
            continue
        qty = device.get('quantity', 1) or 1
        row = format_device_for_export(device, qty)
        if row:
            rows.append(row)
    return rows


def export_excel(devices, scenario, budget, matched_rules=None, floorplan='', budget_tier=''):
    """
    生成 Excel (.xlsx) 购物清单（双 Sheet 美化版）

    参数:
        devices: 设备列表 (list[dict])
        scenario: 场景/方案名称 (str)
        budget: 预算 (int)
        matched_rules: 匹配的联动规则列表（可选，用于 Sheet2 使用说明）
        floorplan: 户型名称（可选，用于 Sheet2）
        budget_tier: 预算档位（可选，用于 Sheet2）

    返回:
        bytes: xlsx 文件内容，供 st.download_button 使用
    """
    try:
        rows = _prepare_export_rows(devices)
        total_price = sum(r.get('小计(元)', 0) for r in rows)

        wb = Workbook()

        # ========== Sheet1: 📋 购物清单 ==========
        ws1 = wb.active
        ws1.title = "📋 购物清单"

        # 样式定义
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="微软雅黑", size=10)
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        price_alignment = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        # 标题行
        col_count = len(EXPORT_COLUMNS)
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_cell = ws1.cell(row=1, column=1, value=f"🏠 HomeWizard 智能家居购物清单 — {scenario}")
        title_cell.font = Font(name="微软雅黑", bold=True, size=14, color="1A1A2E")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 32

        # 预算信息行
        ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        budget_cell = ws1.cell(row=2, column=1, value=f"💰 预算上限：¥{budget:,.2f}  |  生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
        budget_cell.font = Font(name="微软雅黑", size=10, color="64748B")
        budget_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[2].height = 22

        # 表头行（第 4 行）
        header_row = 4
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
            cell = ws1.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws1.row_dimensions[header_row].height = 28

        # 数据行
        for idx, row_data in enumerate(rows, 1):
            row_num = header_row + idx
            for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
                value = row_data.get(col_name, '')
                cell = ws1.cell(row=row_num, column=col_idx, value=value)
                cell.font = cell_font
                cell.border = thin_border
                # 价格列：带 ¥ 格式，2 位小数
                if col_name in ('单价(元)', '小计(元)'):
                    cell.alignment = price_alignment
                    cell.number_format = "¥#,##0.00"
                elif col_name == '数量':
                    cell.alignment = center_alignment
                elif col_name == '购买链接':
                    # 超链接格式
                    if value and str(value).startswith('http'):
                        cell.hyperlink = str(value)
                        cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
                else:
                    cell.alignment = cell_alignment
            ws1.row_dimensions[row_num].height = 24

        # 总价行
        total_row = header_row + len(rows) + 1
        # 合并前 4 列作为标签
        ws1.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
        total_label = ws1.cell(row=total_row, column=1, value="💰 总计")
        total_label.font = Font(name="微软雅黑", bold=True, size=12, color="1A1A2E")
        total_label.alignment = Alignment(horizontal="right", vertical="center")
        total_label.fill = total_fill
        total_label.border = thin_border

        total_value = ws1.cell(row=total_row, column=6, value=total_price)
        total_value.font = Font(name="微软雅黑", bold=True, size=12, color="DC2626")
        total_value.alignment = price_alignment
        total_value.number_format = "¥#,##0.00"
        total_value.fill = total_fill
        total_value.border = thin_border

        # 剩余列填充边框
        for col_idx in range(7, col_count + 1):
            cell = ws1.cell(row=total_row, column=col_idx, value='')
            cell.fill = total_fill
            cell.border = thin_border
        ws1.row_dimensions[total_row].height = 28

        # 列宽设置
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
            width = COLUMN_WIDTHS.get(col_name, 15)
            ws1.column_dimensions[get_column_letter(col_idx)].width = width

        # 冻结表头（前 4 行）
        ws1.freeze_panes = "A5"

        # ========== Sheet2: 📖 使用说明 ==========
        ws2 = wb.create_sheet("📖 使用说明")

        info_font = Font(name="微软雅黑", bold=True, size=12, color="1A1A2E")
        content_font = Font(name="微软雅黑", size=10, color="333333")
        section_font = Font(name="微软雅黑", bold=True, size=11, color="667EEA")

        # 标题
        ws2.merge_cells("A1:D1")
        t = ws2.cell(row=1, column=1, value="📖 HomeWizard 方案使用说明")
        t.font = Font(name="微软雅黑", bold=True, size=14, color="1A1A2E")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 30

        # 基本信息
        info_items = [
            ("🕐 生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("🏠 方案名称", scenario or '智能方案'),
            ("📐 户型", floorplan or '未指定'),
            ("📊 预算档位", budget_tier or '未指定'),
            ("💰 预算上限", f"¥{budget:,.2f}"),
            ("📦 设备总数", f"{len(rows)} 件"),
            ("💵 方案总价", f"¥{total_price:,.2f}"),
        ]
        for idx, (label, value) in enumerate(info_items, 3):
            ws2.cell(row=idx, column=1, value=label).font = info_font
            ws2.cell(row=idx, column=1).alignment = Alignment(horizontal="right", vertical="center")
            ws2.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)
            c = ws2.cell(row=idx, column=2, value=value)
            c.font = content_font
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws2.row_dimensions[idx].height = 22

        # 联动规则区
        rule_start_row = len(info_items) + 5
        ws2.merge_cells(start_row=rule_start_row, start_column=1, end_row=rule_start_row, end_column=4)
        rule_title = ws2.cell(row=rule_start_row, column=1, value="🔗 联动规则列表")
        rule_title.font = section_font
        rule_title.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[rule_start_row].height = 26

        if matched_rules:
            for ri, rule in enumerate(matched_rules[:10]):
                row_num = rule_start_row + 1 + ri
                rname = rule.get('name', '')
                rdesc = rule.get('description', '')
                rate = int(rule.get('match_rate', 0) * 100)
                actions = '；'.join(rule.get('actions', [])[:2]) if rule.get('actions') else ''
                ws2.cell(row=row_num, column=1, value=f"  {rule.get('icon', '🔗')}").font = content_font
                ws2.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
                ws2.cell(row=row_num, column=2, value=f"{rname}（匹配 {rate}%）").font = content_font
                ws2.cell(row=row_num, column=4, value=actions).font = Font(name="微软雅黑", size=9, color="64748B")
                ws2.row_dimensions[row_num].height = 20
        else:
            ws2.merge_cells(start_row=rule_start_row + 1, start_column=1, end_row=rule_start_row + 1, end_column=4)
            ws2.cell(row=rule_start_row + 1, column=1, value="  暂无匹配的联动规则").font = Font(name="微软雅黑", size=10, color="94A3B8")

        # 注意事项
        note_start = rule_start_row + (len(matched_rules[:10]) if matched_rules else 1) + 3
        ws2.merge_cells(start_row=note_start, start_column=1, end_row=note_start, end_column=4)
        note_title = ws2.cell(row=note_start, column=1, value="⚠️ 注意事项")
        note_title.font = section_font
        note_title.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[note_start].height = 26

        notes = [
            "1. 购买链接为京东搜索直达，可自行比价后下单",
            "2. Zigbee 设备需搭配网关使用，Wi-Fi 设备直连路由器",
            "3. 开关类设备请确认零线是否到位（老房可能只有火线）",
            "4. 建议先购入网关+1~2 个设备试用，确认稳定性后再批量采购",
            "5. 所有设备均支持本地控制，断网仍可基础操作",
            "6. 本清单由 HomeWizard 自动生成，仅供参考",
        ]
        for ni, note in enumerate(notes):
            row_num = note_start + 1 + ni
            ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            c = ws2.cell(row=row_num, column=1, value=note)
            c.font = content_font
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws2.row_dimensions[row_num].height = 20

        # Sheet2 列宽
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 20
        ws2.column_dimensions['D'].width = 30

        # 转为 bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception:
        # 所有异常捕获，返回 None 由上层处理
        return None


def export_csv(devices, scenario, budget, matched_rules=None, floorplan='', budget_tier=''):
    """
    生成纯文本 CSV 购物清单（UTF-8 BOM，Excel 打开中文不乱码）
    包含与 Excel 相同的 13 列字段。

    参数:
        devices: 设备列表 (list[dict])
        scenario: 场景/方案名称 (str)
        budget: 预算 (int)
        matched_rules: 匹配的联动规则列表（可选）
        floorplan: 户型名称（可选）
        budget_tier: 预算档位（可选）

    返回:
        str: CSV 文本内容
    """
    try:
        output = io.StringIO()
        # UTF-8 BOM 确保 Excel 正确显示中文
        output.write("\ufeff")
        writer = csv.writer(output)

        # 标题信息
        writer.writerow([f"🏠 HomeWizard 智能家居购物清单 — {scenario}"])
        writer.writerow([f"💰 预算上限：¥{budget:.2f}  |  生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        if floorplan:
            writer.writerow([f"📐 户型：{floorplan}  |  📊 预算档位：{budget_tier or '未指定'}"])
        writer.writerow([])

        # 表头
        writer.writerow(EXPORT_COLUMNS)

        # 数据行
        rows = _prepare_export_rows(devices)
        total_price = 0
        for row_data in rows:
            total_price += row_data.get('小计(元)', 0)
            writer.writerow([row_data.get(col, '') for col in EXPORT_COLUMNS])

        # 总价行
        writer.writerow([])
        writer.writerow(['', '', '', '', '💰 总计', f"¥{total_price:.2f}", '', '', '', '', '', '', ''])

        # 联动规则区
        if matched_rules:
            writer.writerow([])
            writer.writerow(['🔗 联动规则列表'])
            for rule in matched_rules[:10]:
                rname = rule.get('name', '')
                rate = int(rule.get('match_rate', 0) * 100)
                actions = '；'.join(rule.get('actions', [])[:2]) if rule.get('actions') else ''
                writer.writerow([f"{rule.get('icon', '🔗')} {rname}（匹配 {rate}%）", '', '', '', '', '', '', '', '', '', '', '', actions])

        return output.getvalue()
    except Exception:
        return ''
