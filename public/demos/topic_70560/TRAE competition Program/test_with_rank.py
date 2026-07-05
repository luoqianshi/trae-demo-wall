import requests
import openpyxl
import os

TEST_DIR = "test_data"
os.makedirs(TEST_DIR, exist_ok=True)

def create_test_file(filename, data, headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in data:
        ws.append(row)
    filepath = os.path.join(TEST_DIR, filename)
    wb.save(filepath)
    return filepath

headers1 = ["班级", "姓名", "学号", "语文", "数学", "英语", "总分", "排名"]
headers2 = ["班级", "姓名", "学号", "语文", "数学", "英语"]

term1_data = [
    ["初一1班", "张三", "2026001", 85, 92, 88, 265, 5],
    ["初一1班", "李四", "2026002", 78, 85, 90, 253, 8],
    ["初一1班", "王五", "2026003", 90, 88, 92, 270, 3],
]

term2_data = [
    ["初一1班", "张三", "2026001", 88, 90, 92],
    ["初一1班", "李四", "2026002", 82, 88, 85],
    ["初一1班", "赵六", "2026004", 75, 80, 85],
]

file1 = create_test_file("第1学期成绩.xlsx", term1_data, headers1)
file2 = create_test_file("第2学期成绩.xlsx", term2_data, headers2)

print("测试文件创建成功")
print(f"文件1: {file1} (含排名)")
print(f"文件2: {file2} (无排名)")
print()

url = "http://localhost:8000/api/data/clean_and_merge"
files = [
    ('files', open(file1, 'rb')),
    ('files', open(file2, 'rb'))
]

response = requests.post(url, files=files)
print(f"响应状态码: {response.status_code}")
result = response.json()

if response.status_code != 200:
    print(f"错误信息: {result}")
    try:
        os.remove(file1)
        os.remove(file2)
    except:
        pass
    exit(1)

print("\n=== 清洗合并结果 ===")
print(f"总文件数: {result['total_files']}")
print(f"合并后学生数: {result['merged_students']}")

download_url = f"http://localhost:8000/api/data/download/{result['output_filename']}"
download_response = requests.get(download_url)

output_file = os.path.join(TEST_DIR, "merged_with_rank.xlsx")
with open(output_file, 'wb') as f:
    f.write(download_response.content)

print(f"\n合并结果已下载: {output_file}")

wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("\n合并后表格内容:")
headers = []
for row in ws.iter_rows(values_only=True):
    if not headers:
        headers = list(row)
        print(f"表头: {headers}")
        print("-" * 200)
    else:
        print(row)

print("\n=== 预期格式 ===")
print("表头顺序: 班级, 姓名, 学号, 数据来源, 第1学期_语文, 第1学期_数学, 第1学期_英语, 第1学期_总分, 第1学期_排名, 第2学期_语文, 第2学期_数学, 第2学期_英语")
print("同一学期的同一学科所有字段连续显示")

try:
    os.remove(file1)
    os.remove(file2)
    os.remove(output_file)
except:
    pass