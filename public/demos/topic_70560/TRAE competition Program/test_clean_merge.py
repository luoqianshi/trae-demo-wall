import requests
import openpyxl
import os

TEST_DIR = "test_data"
os.makedirs(TEST_DIR, exist_ok=True)

def create_test_file(filename, data, headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in data:
        ws.append(row)
    filepath = os.path.join(TEST_DIR, filename)
    wb.save(filepath)
    return filepath

headers = ["班级", "姓名", "学号", "语文", "数学", "英语"]

term1_data = [
    ["初一1班", "张三", "2026001", 85, 92, 88],
    ["初一1班", "李四", "2026002", 78, 85, 90],
    ["初一1班", "王五", "2026003", 90, 88, 92],
    ["初一1班", "张三", "2026001", 85, 92, 88],
]

term2_data = [
    ["初一1班", "张三", "2026001", 88, 90, 92],
    ["初一1班", "李四", "2026002", 82, 88, 85],
    ["初一1班", "赵六", "2026004", 75, 80, 85],
]

term3_data = [
    ["初一1班", "张三", "2026001", 90, 95, 89],
    ["初一1班", "李四", "2026002", 85, 90, 88],
    ["初一1班", "王五", "2026003", 88, 92, 90],
    ["初一1班", "赵六", "2026004", 80, 85, 88],
]

file1 = create_test_file("第1学期成绩.xlsx", term1_data, headers)
file2 = create_test_file("第2学期成绩.xlsx", term2_data, headers)
file3 = create_test_file("第3学期成绩.xlsx", term3_data, headers)

print("测试文件创建成功")
print(f"文件1: {file1} (含重复行)")
print(f"文件2: {file2}")
print(f"文件3: {file3}")
print()

url = "http://localhost:8000/api/data/clean_and_merge"
files = [
    ('files', open(file1, 'rb')),
    ('files', open(file2, 'rb')),
    ('files', open(file3, 'rb'))
]

response = requests.post(url, files=files)
print(f"响应状态码: {response.status_code}")
result = response.json()

if response.status_code != 200:
    print(f"错误信息: {result}")
    try:
        os.remove(file1)
        os.remove(file2)
        os.remove(file3)
    except:
        pass
    exit(1)

print("\n=== 清洗合并结果 ===")
print(f"总文件数: {result['total_files']}")
print(f"原始总行数: {result['total_original_rows']}")
print(f"清洗后行数: {result['total_cleaned_rows']}")
print(f"移除行数: {result['total_removed_rows']}")
print(f"合并后学生数: {result['merged_students']}")
print(f"移除重复行数: {result['removed_duplicates']}")

print("\n各文件处理详情:")
for fr in result['file_results']:
    print(f"  {fr['filename']}: 原始{fr['original_rows']}行 -> 清洗后{fr['cleaned_rows']}行 (移除{fr['removed_rows']}行)")

download_url = f"http://localhost:8000/api/data/download/{result['output_filename']}"
download_response = requests.get(download_url)

output_file = os.path.join(TEST_DIR, "cleaned_merged_result.xlsx")
with open(output_file, 'wb') as f:
    f.write(download_response.content)

print(f"\n合并结果已下载: {output_file}")

wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("\n合并后表格内容:")
headers = []
for row in ws.iter_rows(values_only=True):
    if not headers:
        headers = list(row)
        print(f"表头: {headers}")
        print("-" * 150)
    else:
        print(row)

print("\n=== 预期结果 ===")
print("张三: 应包含3个学期的成绩，且第1学期的重复行已被移除")
print("李四: 应包含3个学期的成绩")
print("王五: 应包含第1学期和第3学期的成绩")
print("赵六: 应包含第2学期和第3学期的成绩")
print("总计: 4名学生")

try:
    os.remove(file1)
    os.remove(file2)
    os.remove(file3)
    os.remove(output_file)
except:
    pass