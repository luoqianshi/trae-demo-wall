import os
import re
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import xlrd


def is_xlsx_file(file_path: str) -> bool:
    return file_path.lower().endswith('.xlsx')


def is_xls_file(file_path: str) -> bool:
    return file_path.lower().endswith('.xls')


def read_excel_file(file_path: str) -> List[List[Any]]:
    if is_xlsx_file(file_path):
        return read_xlsx_file(file_path)
    elif is_xls_file(file_path):
        return read_xls_file(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_path}")


def read_xlsx_file(file_path: str) -> List[List[Any]]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    return read_worksheet_data(ws)


def read_xls_file(file_path: str) -> List[List[Any]]:
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    rows = []
    for row_idx in range(ws.nrows):
        row = []
        for col_idx in range(ws.ncols):
            cell_value = ws.cell_value(row_idx, col_idx)
            row.append(cell_value)
        rows.append(row)
    return rows


def read_worksheet_data(ws: Worksheet) -> List[List[Any]]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def write_excel_file(file_path: str, data: List[List[Any]], headers: Optional[List[str]] = None):
    wb = Workbook()
    ws = wb.active
    
    if headers:
        ws.append(headers)
    
    for row in data:
        ws.append(row)
    
    wb.save(file_path)


def get_worksheet_dimensions(data: List[List[Any]]) -> Dict:
    rows = len(data)
    cols = max(len(row) for row in data) if rows > 0 else 0
    return {"rows": rows, "cols": cols}


def is_empty_row(row: List[Any]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def is_valid_header_cell(cell: Any) -> bool:
    if cell is None:
        return False
    str_val = str(cell).strip()
    return len(str_val) > 0 and len(str_val) <= 50


def detect_header_row(data: List[List[Any]]) -> int:
    for idx, row in enumerate(data[:10]):
        if is_empty_row(row):
            continue
        valid_cells = sum(1 for cell in row if is_valid_header_cell(cell))
        if valid_cells >= 2 and valid_cells > len(row) * 0.3:
            return idx
    return 0


def get_column_names(data: List[List[Any]], header_row: int = 0) -> List[str]:
    if header_row >= len(data):
        return []
    return [str(cell).strip() if cell is not None else "" for cell in data[header_row]]


def get_data_rows(data: List[List[Any]], header_row: int = 0) -> List[List[Any]]:
    return data[header_row + 1:]


def normalize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return round(value, 2)
    return str(value).strip()


def normalize_row(row: List[Any]) -> List[Any]:
    return [normalize_cell_value(cell) for cell in row]


def remove_empty_columns(data: List[List[Any]]) -> List[List[Any]]:
    if not data:
        return data
    
    cols = max(len(row) for row in data)
    non_empty_cols = []
    
    for col_idx in range(cols):
        has_value = False
        for row in data:
            if col_idx < len(row) and row[col_idx] is not None and str(row[col_idx]).strip() != "":
                has_value = True
                break
        if has_value:
            non_empty_cols.append(col_idx)
    
    result = []
    for row in data:
        new_row = []
        for col_idx in non_empty_cols:
            if col_idx < len(row):
                new_row.append(row[col_idx])
            else:
                new_row.append("")
        result.append(new_row)
    
    return result


def remove_empty_rows(data: List[List[Any]]) -> List[List[Any]]:
    return [row for row in data if not is_empty_row(row)]


def find_column_index(headers: List[str], target: str, threshold: int = 80) -> int:
    from fuzzywuzzy import fuzz
    
    for idx, header in enumerate(headers):
        score = fuzz.partial_ratio(str(header).lower(), target.lower())
        if score >= threshold:
            return idx
    return -1


def extract_column(data: List[List[Any]], column_index: int) -> List[Any]:
    return [row[column_index] if column_index < len(row) else "" for row in data]


def validate_score(value: Any) -> bool:
    try:
        float_val = float(value)
        return 0 <= float_val <= 150
    except (ValueError, TypeError):
        return False


def is_numeric_column(data: List[Any]) -> bool:
    numeric_count = sum(1 for val in data if val is not None and str(val).strip() != "" and validate_score(val))
    return numeric_count > len(data) * 0.5