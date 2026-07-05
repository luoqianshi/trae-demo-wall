import requests
import openpyxl
import os

TEST_DIR = "test_data"
os.makedirs(TEST_DIR, exist_ok=True)

def create_test_file(filename, data, headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in data:
        ws.append(row)
    filepath = os.path.join(TEST_DIR, filename)
    wb.save(filepath)
    return filepath

headers = ["班级", "姓名", "学号", "语文", "数学", "英语"]

data = [
    ["初一1班", "张三", "2026001", 85, 92, 88],
    ["初一1班", "李四", "2026002", 78, 85, 90],
    ["初一2班", "王五", "2026003", 90, 88, 92],
    ["初一1班", "张三", "2026001", 85, 92, 88],
    ["初一1班", "张三", "2026001", 85, 92, 88],
    ["初一2班", "赵六", "2026004", 75, 80, 85],
    ["初二1班", "孙七", "2026005", 88, 85, 90],
    ["初二1班", "孙七", "2026005", 88, 85, 90],
]

file1 = create_test_file("clean_test.xlsx", data, headers)

print("测试文件创建成功")
print(f"文件: {file1}")
print(f"原始数据行数: {len(data)}")
print()
print("数据内容:")
print(f"{'班级':<10} {'姓名':<8} {'学号':<10} {'语文':<5} {'数学':<5} {'英语':<5}")
print("-" * 50)
for row in data:
    print(f"{row[0]:<10} {row[1]:<8} {row[2]:<10} {row[3]:<5} {row[4]:<5} {row[5]:<5}")

url = "http://localhost:8000/api/data/clean"
files = {'file': open(file1, 'rb')}

response = requests.post(url, files=files)
print(f"\n响应状态码: {response.status_code}")
result = response.json()

if response.status_code != 200:
    print(f"错误信息: {result}")
    try:
        os.remove(file1)
    except:
        pass
    exit(1)

print("\n=== 清洗结果 ===")
print(f"原始行数: {result['original_rows']}")
print(f"清洗后行数: {result['cleaned_rows']}")
print(f"移除行数: {result['removed_rows']}")
print(f"重复行数: {result['duplicate_count']}")

if result['duplicate_count'] > 0:
    print("\n重复行详情:")
    for idx, dup in enumerate(result['duplicate_details']):
        print(f"\n重复 #{idx+1}:")
        print(f"  行号: 第{dup['row_index'] + 2}行")
        print(f"  数据: {', '.join([f'{k}: {v}' for k, v in dup.items() if k != 'row_index'])}")

download_url = f"http://localhost:8000/api/data/download/{result['output_filename']}"
download_response = requests.get(download_url)

output_file = os.path.join(TEST_DIR, "cleaned_result.xlsx")
with open(output_file, 'wb') as f:
    f.write(download_response.content)

print(f"\n清洗结果已下载: {output_file}")

wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("\n清洗后表格内容:")
for row in ws.iter_rows(values_only=True):
    print(row)

row_count = ws.max_row - 1
expected = 5
if row_count == expected:
    print(f"\n✓ 清洗成功！预期{expected}人，实际{row_count}人")
else:
    print(f"\n✗ 清洗结果异常，预期{expected}人，实际{row_count}人")

try:
    os.remove(file1)
except:
    pass
try:
    os.remove(output_file)
except:
    pass