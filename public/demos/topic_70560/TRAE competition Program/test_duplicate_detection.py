import requests
import openpyxl
import os

TEST_DIR = "test_data"
os.makedirs(TEST_DIR, exist_ok=True)

def create_test_file(filename, data, headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.append(headers)
    for row in data:
        ws.append(row)
    
    filepath = os.path.join(TEST_DIR, filename)
    wb.save(filepath)
    return filepath

headers = ["班级", "姓名", "学号", "语文", "数学", "英语"]

data1 = [
    ["初一1班", "张三", "2026001", 85, 92, 88],
    ["初一1班", "李四", "2026002", 78, 85, 90],
    ["初一2班", "王五", "2026003", 90, 88, 92],
    ["初一1班", "张三", "2026001", 85, 92, 88],
    ["初一2班", "张三", "2026001", 82, 79, 85],
]

data2 = [
    ["初二1班", "赵六", "", 60, 50, 40],
    ["初二1班", "赵六", "", 60, 50, 40],
    ["初二1班", "钱七", "2026005", -10, 77, 66],
    ["初二1班", "孙八", "2026006", 99, 95, 98],
    ["初二2班", "孙八", "2026006", 99, 95, 98],
]

file1 = create_test_file("duplicate_test1.xlsx", data1, headers)
file2 = create_test_file("duplicate_test2.xlsx", data2, headers)

print("测试文件创建成功")
print(f"文件1: {file1}")
print(f"文件2: {file2}")
print(f"文件1数据行数: {len(data1)}")
print(f"文件2数据行数: {len(data2)}")
print()
print("文件1数据:")
print(f"{'班级':<10} {'姓名':<8} {'学号':<10} {'语文':<5} {'数学':<5} {'英语':<5}")
print("-" * 50)
for row in data1:
    print(f"{row[0]:<10} {row[1]:<8} {row[2]:<10} {row[3]:<5} {row[4]:<5} {row[5]:<5}")

print("\n文件2数据:")
print(f"{'班级':<10} {'姓名':<8} {'学号':<10} {'语文':<5} {'数学':<5} {'英语':<5}")
print("-" * 50)
for row in data2:
    print(f"{row[0]:<10} {row[1]:<8} {row[2]:<10} {row[3]:<5} {row[4]:<5} {row[5]:<5}")

url = "http://localhost:8000/api/match/merge"
files = [
    ('files', open(file1, 'rb')),
    ('files', open(file2, 'rb')),
]

response = requests.post(url, files=files)
print(f"\n响应状态码: {response.status_code}")
result = response.json()

if response.status_code != 200:
    print(f"错误信息: {result}")
    try:
        os.remove(file1)
    except:
        pass
    try:
        os.remove(file2)
    except:
        pass
    exit(1)

print("\n=== 合并结果 ===")
print(f"原始学生数: {result['total_source_students']}")
print(f"合并后学生数: {result['merged_students']}")
print(f"移除的重复行数: {result['removed_duplicates']}")

print("\n=== 重复检测详情 ===")
duplicate_info = result['duplicate_info']
print(f"完全重复行数: {duplicate_info['exact_duplicates']}")

if duplicate_info['exact_duplicates'] > 0:
    print("\n完全重复行详情:")
    for idx, dup in enumerate(duplicate_info['exact_duplicate_details']):
        print(f"\n重复 #{idx+1}:")
        print(f"  原始行: {dup['original']['name']} - {dup['original']['class']} - {dup['original']['source']} 第{dup['original']['row_index']+1}行")
        print(f"  重复行: {dup['duplicate']['name']} - {dup['duplicate']['class']} - {dup['duplicate']['source']} 第{dup['duplicate']['row_index']+1}行")

download_url = f"http://localhost:8000/api/match/download/{result['output_filename']}"
download_response = requests.get(download_url)

output_file = os.path.join(TEST_DIR, "merged_result.xlsx")
with open(output_file, 'wb') as f:
    f.write(download_response.content)

print(f"\n合并结果已下载: {output_file}")

wb = openpyxl.load_workbook(output_file)
ws = wb.active

print("\n合并结果表格:")
for row in ws.iter_rows(values_only=True):
    print(row)

row_count = ws.max_row - 1
print(f"\n合并后学生总数: {row_count}")
print(f"预期学生总数: 8 (初一1班张三、初一1班李四、初一2班王五、初一2班张三、初二1班赵六、初二1班钱七、初二1班孙八、初二2班孙八)")
print("说明：孙八在初二1班和初二2班是不同班级的记录，应保持独立；初一2班的张三与初一1班的张三名相同但班级不同，应保持独立")

expected = 8
if row_count == expected:
    print("✓ 合并成功！完全重复的行已去重，不同班级的同名学生保持独立")
else:
    print(f"✗ 合并结果异常，预期{expected}人，实际{row_count}人")

try:
    os.remove(file1)
except:
    pass
try:
    os.remove(file2)
except:
    pass
try:
    os.remove(output_file)
except:
    pass