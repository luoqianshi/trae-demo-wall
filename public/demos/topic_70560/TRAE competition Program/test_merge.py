import requests
import openpyxl
import os

TEST_DIR = "test_data"
os.makedirs(TEST_DIR, exist_ok=True)

def create_test_file(filename, data, headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.append(headers)
    for row in data:
        ws.append(row)
    
    filepath = os.path.join(TEST_DIR, filename)
    wb.save(filepath)
    return filepath

headers = ["序号", "姓名", "班级", "学号", "语文", "数学", "英语"]

data1 = [
    [1, "张三", "初一1班", "2024001", 85, 92, 88],
    [2, "李四", "初一1班", "2024002", 78, 85, 90],
    [3, "王五", "初一2班", "2024003", 90, 88, 92],
    [4, "赵六", "初一2班", "2024004", 75, 80, 85],
]

data2 = [
    [1, "张三", "初一1班", "2024001", 88, 95, 90],
    [2, "李四", "初一1班", "2024002", 82, 88, 92],
    [3, "王五", "初一2班", "2024003", 89, 90, 91],
    [5, "孙七", "初一3班", "2024005", 92, 93, 94],
]

file1 = create_test_file("exam1.xlsx", data1, headers)
file2 = create_test_file("exam2.xlsx", data2, headers)

print("测试文件创建成功")
print(f"文件1: {file1}")
print(f"文件2: {file2}")

url = "http://localhost:8000/api/match/merge"
files = [
    ('files', open(file1, 'rb')),
    ('files', open(file2, 'rb'))
]

response = requests.post(url, files=files)
print(f"\n响应状态码: {response.status_code}")
result = response.json()
print(f"响应内容: {result}")

if result.get("success"):
    download_url = f"http://localhost:8000/api/match/download/{result['output_filename']}"
    download_response = requests.get(download_url)
    
    output_file = os.path.join(TEST_DIR, "merged_result.xlsx")
    with open(output_file, 'wb') as f:
        f.write(download_response.content)
    
    print(f"\n合并结果已下载: {output_file}")
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    print("\n合并结果表格:")
    for row in ws.iter_rows(values_only=True):
        print(row)
    
    row_count = ws.max_row - 1
    print(f"\n合并后学生总数: {row_count}")
    print(f"预期学生总数: 5 (张三、李四、王五、赵六、孙七)")
    
    if row_count == 5:
        print("✓ 合并成功！每个学生只出现一次")
    else:
        print(f"✗ 合并结果异常，预期5人，实际{row_count}人")

os.remove(file1)
os.remove(file2)
if os.path.exists(output_file):
    os.remove(output_file)