# v2.4.0-stable Modular Pipeline Component
import os
import sys
import json
import csv
import re
import shutil
import glob
import traceback
import time
from pathlib import Path
from datetime import datetime

from pipeline_modules.constants import *

# Optional dependencies
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

def safe_print(message: object = "") -> None:
    text = str(message)
    try:
        print(text)
    except UnicodeEncodeError:
        try:
            print(text.encode("utf-8", errors="replace").decode("utf-8", errors="replace"))
        except Exception:
            pass

class ExtractorMixin:
    def extract(self, incremental=False):
        print(f"[*] 开始提取: {self.sku_name} (incremental={incremental})")
        print(f"[*] 沙箱: {self.scratch_dir}")

        if not incremental:
            # 强行物理清除 outfit_images_cleared.flag，保证全新一轮生图被清空！
            flag_path = os.path.join(self.scratch_dir, 'outfit_images_cleared.flag')
            if os.path.exists(flag_path):
                try:
                    os.remove(flag_path)
                    print("[*] 全新提取物料，物理清除旧生图标志锁")
                except Exception:
                    pass

        # 0) 先行提取 Excel 里的所有图片到 excel_images 目录 (WPS DISPIMG 支持)
        self._extract_images_from_all_xlsx()

        old_state = {}
        if incremental:
            state_path = os.path.join(self.scratch_dir, 'asset_state.json')
            if os.path.exists(state_path):
                try:
                    with open(state_path, 'r', encoding='utf-8') as f:
                        old_state = json.load(f)
                    print(f"[*] 已载入物理资产状态缓存，共 {len(old_state)} 项")
                except Exception as e:
                    print(f"[!] 载入资产缓存失败: {e}")

        # 需要跳过的目录
        skip_dirs = {'.scratch', '自动生成素材', 'scratch', 'parts', '.git', '__pycache__'}
        # 需要跳过的文件模式
        sku_prefix = self.sku_name.split('-')[0] if '-' in self.sku_name else self.sku_name

        def _should_skip_file(fname):
            """跳过历史生成文件，避免上下文膨胀"""
            fl = fname.lower()
            # 跳过上次生成的手册
            if fname.startswith(sku_prefix) and fname.endswith('.md'):
                return True
            if fl.endswith('.review.md'):
                return True
            if fl.endswith(('.tmp', '.bak')):
                return True
            return False

        # 1) 递归扫描
        manifest_entries = []
        for root, dirs, files in os.walk(self.target_dir):
            # 跳过特定目录
            dirs[:] = [d for d in dirs if d not in skip_dirs]
            for fname in files:
                if _should_skip_file(fname):
                    continue
                fpath = os.path.join(root, fname)
                ext = os.path.splitext(fname)[1].lower()
                rel = os.path.relpath(fpath, self.target_dir)
                rel_clean = rel.replace('\\', '/')

                # 动态跳过位于 stable_assets_dir 中的历史生成文件，避免重复扫描和注入
                if self.stable_assets_rel in rel_clean:
                    continue

                entry = {
                    'relative_path': rel_clean,
                    'absolute_path': fpath.replace('\\', '/'),
                    'filename': fname,
                    'extension': ext,
                    'file_type': self._classify_file(ext),
                    'size_bytes': os.path.getsize(fpath),
                    'mtime': os.path.getmtime(fpath),
                    'extraction_status': 'pending',
                    'error_message': None,
                }
                # 图片额外字段
                if ext in IMAGE_EXTS:
                    reused = False
                    if incremental and old_state and rel_clean in old_state:
                        old_entry = old_state[rel_clean]
                        # 检查物理属性是否完全未变
                        cur_mtime_ns = os.stat(fpath).st_mtime_ns
                        if old_entry.get('size_bytes') == entry['size_bytes'] and old_entry.get('mtime_ns') == cur_mtime_ns:
                            # 继承历史分析属性
                            for k in ('width', 'height', 'aspect_ratio', 'path_hints', 'system_flags', 'needs_review_reason'):
                                if k in old_entry:
                                    entry[k] = old_entry[k]
                            entry['extraction_status'] = 'ok'
                            reused = True
                            print(f"    [增量复用] 已从资产库秒级复用图片属性: {rel_clean}")

                    if not reused:
                        entry.update(self._analyze_image(fpath, rel))
                manifest_entries.append(entry)

        # 2) 提取文本内容
        dump_lines = [f'=== 商品资料提取: {self.sku_name} ===',
                      f'=== 提取时间: {datetime.now().isoformat()} ===\n']
        new_entries = []
        for e in manifest_entries:
            ext = e['extension']
            fpath = e['absolute_path']
            try:
                if ext in TEXT_EXTS:
                    content = self._read_text(fpath)
                    dump_lines.append(f'\n--- [文本] {e["relative_path"]} ---')
                    dump_lines.append(content)
                    e['extraction_status'] = 'ok'
                    e['extracted_text_length'] = len(content)
                elif ext == '.pdf':
                    txt, thumb = self._extract_pdf(fpath)
                    dump_lines.append(f'\n--- [PDF] {e["relative_path"]} ---')
                    dump_lines.append(txt if txt else '(无文本内容)')
                    e['extraction_status'] = 'ok'
                    e['extracted_text_length'] = len(txt) if txt else 0
                    if thumb:
                        e['thumbnail_stable_path'] = thumb.replace('\\','/')
                        # 动态注入一个独立的图片 entry 放入新 entries 候选池
                        try:
                            img_rel = os.path.relpath(thumb, self.target_dir).replace('\\', '/')
                            img_entry = {
                                'relative_path': img_rel,
                                'absolute_path': thumb.replace('\\', '/'),
                                'filename': os.path.basename(thumb),
                                'extension': '.png',
                                'file_type': 'image',
                                'size_bytes': os.path.getsize(thumb),
                                'mtime': os.path.getmtime(thumb),
                                'extraction_status': 'ok',
                                'error_message': None,
                            }
                            img_entry.update(self._analyze_image(thumb, img_rel))
                            new_entries.append(img_entry)
                            print(f"    [注入] 已为 PDF 缩略图注册独立图片 Entry: {img_rel}")
                        except Exception as img_ex:
                            print(f"    [警告] PDF 缩略图注册独立图片 Entry 失败: {img_ex}")
                elif ext == '.docx':
                    txt = self._extract_docx(fpath)
                    dump_lines.append(f'\n--- [Word] {e["relative_path"]} ---')
                    dump_lines.append(txt if txt else '(无文本内容)')
                    e['extraction_status'] = 'ok'
                    e['extracted_text_length'] = len(txt) if txt else 0
                elif ext in SPREADSHEET_EXTS:
                    txt, meta = self._extract_spreadsheet(fpath, ext)
                    dump_lines.append(f'\n--- [表格] {e["relative_path"]} ---')
                    dump_lines.append(txt)
                    e['extraction_status'] = 'ok'
                    e['extracted_text_length'] = len(txt)
                    if meta:
                        e.update(meta)
                elif ext in IMAGE_EXTS:
                    e['extraction_status'] = 'ok'
                elif ext in VIDEO_EXTS:
                    e['extraction_status'] = 'skipped_video'
                else:
                    e['extraction_status'] = 'skipped_unknown'
            except Exception as ex:
                e['extraction_status'] = 'failed'
                e['error_message'] = str(ex)
                dump_lines.append(f'\n--- [错误] {e["relative_path"]} ---')
                dump_lines.append(f'提取失败: {ex}')

        if new_entries:
            manifest_entries.extend(new_entries)

        # 3) 关键字段摘要
        dump_lines.append('\n\n=== 关键字段摘要 ===')
        full_text = '\n'.join(dump_lines)
        for kw in KEYWORDS:
            hits = [l.strip() for l in full_text.split('\n') if kw in l][:3]
            if hits:
                dump_lines.append(f'\n[{kw}]:')
                for h in hits:
                    dump_lines.append(f'  {h[:200]}')

        # 4) 写入文件
        dump_path = os.path.join(self.scratch_dir, 'dump.txt')
        with open(dump_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(dump_lines))

        manifest_path = os.path.join(self.scratch_dir, 'manifest.json')
        with open(manifest_path, 'w', encoding='utf-8') as f:
            json.dump(manifest_entries, f, ensure_ascii=False, indent=2)

        manifest_md = self._build_manifest_md(manifest_entries)
        md_path = os.path.join(self.scratch_dir, 'manifest.md')
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(manifest_md)

        # 统计
        stats = self._count_stats(manifest_entries)
        print(f"[+] 提取完成!")
        print(f"    文件总数: {stats['total']}")
        print(f"    图片: {stats['images']} | 文本: {stats['texts']} | 表格: {stats['sheets']}")
        print(f"    PDF/Word: {stats['docs']} | 视频: {stats['videos']}")
        print(f"    失败: {stats['failed']} | 疑似1:1: {stats['square']} | 疑似拼接: {stats['collage']}")
        print(f"    输出: {dump_path}")
        return manifest_entries

    def _classify_file(self, ext):
        if ext in IMAGE_EXTS: return 'image'
        if ext in TEXT_EXTS: return 'text'
        if ext in SPREADSHEET_EXTS: return 'spreadsheet'
        if ext in DOC_EXTS: return 'document'
        if ext in VIDEO_EXTS: return 'video'
        return 'other'

    def _read_text(self, fpath):
        for enc in ['utf-8', 'utf-8-sig', 'gbk', 'gb18030', 'latin-1']:
            try:
                with open(fpath, 'r', encoding=enc) as f:
                    return f.read()
            except (UnicodeDecodeError, UnicodeError):
                continue
        return '(编码无法识别)'

    def _analyze_image(self, fpath, rel_path):
        info = {
            'width': None, 'height': None, 'aspect_ratio': None,
            'path_hints': [],
            'system_flags': [],
            'needs_review_reason': ''
        }
        # 1:1 文件夹检测
        rel_lower = rel_path.replace('\\','/').lower()
        if '1:1' in rel_lower or '1：1' in rel_lower:
            info['system_flags'].append('from_1x1_folder')

        # 不稳定路径检测
        unstable_keywords = ['.uploads', '.trae', 'scratch', 'tmp', 'temp', 'cache', 'clipboard', 'pasted', 'blob']
        if any(k in rel_lower for k in unstable_keywords):
            info['system_flags'].append('temp_or_unstable_path')
            info['needs_review_reason'] = '路径包含临时或不稳定目录，必须复制到商品正式目录下再引用'

        # 文件名 prompt/AI 生成提示标记（不排除，标记为 REVIEW_HINT）
        fname_lower = os.path.basename(rel_path).lower()
        prompt_like_keywords = ['prompt', 'midjourney', 'generated', 'ai_gen', 'dalle', 'stable_diffusion']
        if any(k in fname_lower for k in prompt_like_keywords):
            info['system_flags'].append('filename_prompt_like')

        # 路径提示
        if '棚拍' in rel_lower: info['path_hints'].append('棚拍')
        if '外景' in rel_lower: info['path_hints'].append('外景')
        if '搭配' in rel_lower: info['path_hints'].append('搭配')
        if '物料' in rel_lower or '设计' in rel_lower or '灵感' in rel_lower: info['path_hints'].append('设计物料/灵感')
        # 颜色相关：包含颜色/色卡/color/colour等
        COLOR_PATH_KEYWORDS = ['颜色', '色卡', 'color', 'colour']
        COLOR_NAME_KEYWORDS = ['黑', '白', '灰', '藏青', '藏蓝', '咖', '棕', '米', '红', '绿', '蓝', '紫', '黄',
                               '卡其', '驼', '杏', '粉', '橙', '酒红', '墨绿', '深蓝', '浅灰', '奶白', '焦糖',
                               'black', 'white', 'grey', 'gray', 'navy', 'brown', 'beige', 'red', 'green', 'blue']
        if any(k in rel_lower for k in COLOR_PATH_KEYWORDS) or any(k in fname_lower for k in COLOR_NAME_KEYWORDS):
            info['path_hints'].append('颜色')
        elif '色' in rel_lower:
            info['path_hints'].append('颜色')
        if '面料' in rel_lower or '纹理' in rel_lower: info['path_hints'].append('面料')

        if not HAS_PIL:
            info['needs_review_reason'] = '无法读取图片（Pillow 未安装）'
            return info

        try:
            with Image.open(fpath) as img:
                w, h = img.size
                info['width'] = w
                info['height'] = h
                ratio = round(w / h, 3) if h > 0 else 0
                info['aspect_ratio'] = ratio

                # === 辅助标记逻辑（仅提示，不代替AI决策） ===

                if ratio > 1.8 and w > 2000:
                    info['system_flags'].append('wide_image')
                    info['system_flags'].append('possible_collage')
                    info['needs_review_reason'] = f'宽高比={ratio}, 宽度={w}px，疑似横向拼接'
                elif ratio > 1.5:
                    info['system_flags'].append('wide_image')
                    info['needs_review_reason'] = f'宽高比偏高(ratio={ratio})，需确认是否为拼接图'
                elif 0.9 <= ratio <= 1.1:
                    info['system_flags'].append('square_like')
                    total_px = w * h
                    if total_px > 6_000_000:
                        info['system_flags'].append('possible_collage')
                        info['needs_review_reason'] = f'接近1:1且高分辨率({total_px/1e6:.1f}MP)，疑似2x2网格'
                    else:
                        info['needs_review_reason'] = f'接近1:1，需确认是否为产品静物/色卡或电商主图裁切'
                else:
                    info['system_flags'].append('normal_ratio')

                # 棚拍/外景大图警告
                if ('棚拍' in info['path_hints'] or '外景' in info['path_hints']) and w * h > 4_000_000 and ratio > 1.15:
                    if 'possible_collage' not in info['system_flags']:
                        info['system_flags'].append('possible_collage')
                    reason = f'棚拍/外景大图且偏宽({w}x{h}, ratio={ratio})，疑似多视图'
                    info['needs_review_reason'] = reason if not info['needs_review_reason'] else info['needs_review_reason'] + " | " + reason

                # --- 接缝检测：检查图片在 1/2、1/3 位置是否有垂直/水平接缝 ---
                seam_detected = False
                seam_detail = ''
                try:
                    px = img.convert('RGB').load()
                    for frac_name, frac in [('1/2', 0.5), ('1/3', 1/3), ('2/3', 2/3)]:
                        cx = int(w * frac)
                        if cx <= 0 or cx >= w - 1: continue
                        sample_count = min(20, h // 10)
                        if sample_count < 5: continue
                        diffs = []
                        for sy in range(sample_count):
                            y_pos = int(h * (sy + 1) / (sample_count + 1))
                            left = px[cx - 1, y_pos]
                            right = px[cx + 1, y_pos]
                            diff = sum(abs(left[c] - right[c]) for c in range(3))
                            diffs.append(diff)
                        avg_diff = sum(diffs) / len(diffs)
                        line_vals = [sum(px[cx, int(h * (sy + 1) / (sample_count + 1))]) for sy in range(sample_count)]
                        line_var = sum((v - sum(line_vals)/len(line_vals))**2 for v in line_vals) / len(line_vals)
                        if avg_diff > 80 or (line_var < 50 and avg_diff > 30):
                            seam_detected = True
                            seam_detail = f'垂直接缝@x={frac_name}(差异={avg_diff:.0f})'
                            break

                    if not seam_detected:
                        for frac_name, frac in [('1/2', 0.5), ('1/3', 1/3), ('2/3', 2/3)]:
                            cy = int(h * frac)
                            if cy <= 0 or cy >= h - 1: continue
                            sample_count = min(20, w // 10)
                            if sample_count < 5: continue
                            diffs = []
                            for sx in range(sample_count):
                                x_pos = int(w * (sx + 1) / (sample_count + 1))
                                top = px[x_pos, cy - 1]
                                bot = px[x_pos, cy + 1]
                                diff = sum(abs(top[c] - bot[c]) for c in range(3))
                                diffs.append(diff)
                            avg_diff = sum(diffs) / len(diffs)
                            if avg_diff > 80:
                                seam_detected = True
                                seam_detail = f'水平接缝@y={frac_name}(差异={avg_diff:.0f})'
                                break
                except Exception:
                    pass  # 接缝检测失败不影响后续逻辑

                if seam_detected:
                    if 'possible_collage' not in info['system_flags']:
                        info['system_flags'].append('possible_collage')
                    reason = f'接缝检测: {seam_detail}'
                    info['needs_review_reason'] = reason if not info['needs_review_reason'] else info['needs_review_reason'] + " | " + reason

        except Exception as e:
            info['needs_review_reason'] = f'图片读取异常: {e}'
        return info

    def _extract_pdf(self, fpath):
        try:
            import fitz
        except ImportError:
            return f'(未安装 PyMuPDF，请执行 pip install PyMuPDF)', None
        text_parts = []
        thumb_stable = None
        try:
            doc = fitz.open(fpath)
            for page in doc:
                text_parts.append(page.get_text())
            # 缩略图 → scratch，然后复制到稳定路径
            if len(doc) > 0:
                page0 = doc.load_page(0)
                pix = page0.get_pixmap(matrix=fitz.Matrix(2, 2))
                basename = os.path.splitext(os.path.basename(fpath))[0]
                scratch_thumb = os.path.join(self.scratch_dir, f'{basename}_缩略图.png')
                pix.save(scratch_thumb)
                # 复制到商品目录稳定路径
                os.makedirs(self.stable_assets_dir, exist_ok=True)
                stable_thumb = os.path.join(self.stable_assets_dir, f'{basename}_缩略图.png')
                shutil.copy2(scratch_thumb, stable_thumb)
                thumb_stable = stable_thumb
                print(f"    PDF缩略图: {os.path.relpath(stable_thumb, self.target_dir)}")
            doc.close()
        except Exception as e:
            text_parts.append(f'PDF读取错误: {e}')
        return '\n'.join(text_parts), thumb_stable

    def _extract_docx(self, fpath):
        try:
            import docx
        except ImportError:
            return f'(未安装 python-docx，请执行 pip install python-docx)'
        try:
            doc = docx.Document(fpath)
            return '\n'.join(p.text for p in doc.paragraphs)
        except Exception as e:
            return f'Word读取错误: {e}'

    def _extract_spreadsheet(self, fpath, ext):
        if ext == '.csv':
            return self._extract_csv(fpath)
        elif ext == '.xlsx':
            return self._extract_xlsx(fpath)
        elif ext == '.xls':
            return self._extract_xls(fpath)
        return '(不支持的表格格式)', None

    def _extract_csv(self, fpath):
        for enc in ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']:
            try:
                with open(fpath, 'r', encoding=enc, newline='') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                if not rows:
                    return '(空CSV文件)', None
                lines = [f'CSV文件: {os.path.basename(fpath)}',
                         f'行数: {len(rows)} | 列数: {len(rows[0]) if rows else 0}', '']
                # Markdown 表格
                if rows:
                    header = rows[0]
                    lines.append('| ' + ' | '.join(str(c)[:30] for c in header) + ' |')
                    lines.append('|' + '---|' * len(header))
                    for row in rows[1:50]:  # 最多50行
                        lines.append('| ' + ' | '.join(str(c)[:30] for c in row) + ' |')
                    if len(rows) > 51:
                        lines.append(f'... 共 {len(rows)} 行')
                return '\n'.join(lines), {'extracted_rows': len(rows), 'extracted_cols': len(rows[0]) if rows else 0}
            except (UnicodeDecodeError, UnicodeError):
                continue
        return '(CSV编码无法识别)', None

    def _extract_xlsx(self, fpath):
        if not HAS_OPENPYXL:
            return '(未安装 openpyxl，请执行 pip install openpyxl)', None
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            lines = [f'Excel文件: {os.path.basename(fpath)}',
                     f'Sheet数量: {len(wb.sheetnames)}',
                     f'Sheet列表: {", ".join(wb.sheetnames)}', '']
            meta = {'sheets': [], 'extracted_sheets_count': len(wb.sheetnames)}
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_data = [str(c) if c is not None else '' for c in row]
                    if any(c.strip() for c in row_data):
                        rows.append(row_data)
                lines.append(f'\n### Sheet: {sname}')
                lines.append(f'有效行数: {len(rows)} | 列数: {ws.max_column}')
                if rows:
                    header = rows[0]
                    lines.append('')
                    lines.append('| ' + ' | '.join(c[:30] for c in header) + ' |')
                    lines.append('|' + '---|' * len(header))
                    for row in rows[1:80]:
                        lines.append('| ' + ' | '.join(c[:30] for c in row) + ' |')
                    if len(rows) > 81:
                        lines.append(f'... 共 {len(rows)} 行')
                else:
                    lines.append('(空Sheet)')
                meta['sheets'].append({'name': sname, 'rows': len(rows),
                                        'cols': ws.max_column or 0})
            wb.close()
            return '\n'.join(lines), meta
        except Exception as e:
            return f'Excel读取错误: {e}', None

    def _extract_xls(self, fpath):
        try:
            import xlrd
            wb = xlrd.open_workbook(fpath)
            lines = [f'Excel(.xls)文件: {os.path.basename(fpath)}',
                     f'Sheet数量: {len(wb.sheet_names())}', '']
            for sname in wb.sheet_names():
                ws = wb.sheet_by_name(sname)
                lines.append(f'\n### Sheet: {sname}')
                lines.append(f'行数: {ws.nrows} | 列数: {ws.ncols}')
                for r in range(min(ws.nrows, 80)):
                    vals = [str(ws.cell_value(r, c))[:30] for c in range(ws.ncols)]
                    lines.append('| ' + ' | '.join(vals) + ' |')
            return '\n'.join(lines), None
        except ImportError:
            return '(未安装 xlrd，请执行 pip install xlrd)', None
        except Exception as e:
            return f'.xls读取错误: {e}', None

    def _build_manifest_md(self, entries):
        stats = self._count_stats(entries)
        lines = [f'# 文件清单: {self.sku_name}', '',
                 f'- 文件总数: {stats["total"]}',
                 f'- 图片: {stats["images"]}',
                 f'- 文本: {stats["texts"]}',
                 f'- 表格: {stats["sheets"]}',
                 f'- PDF/Word: {stats["docs"]}',
                 f'- 视频: {stats["videos"]}',
                 f'- 读取失败: {stats["failed"]}',
                 f'- 疑似1:1图: {stats["square"]}',
                 f'- 疑似拼接图: {stats["collage"]}', '']
        failed = [e for e in entries if e['extraction_status'] == 'failed']
        if failed:
            lines.append('## 失败文件')
            for e in failed:
                lines.append(f'- `{e["relative_path"]}`: {e["error_message"]}')
            lines.append('')
        lines.append('## 完整文件列表')
        lines.append('| 路径 | 类型 | 大小 | 状态 | 建议分类 |')
        lines.append('|:---|:---|:---|:---|:---|')
        for e in entries:
            sz = f'{e["size_bytes"]/1024:.1f}KB'
            bucket = e.get("suggested_bucket", "")
            if bucket and "DISCARD" in bucket:
                bucket = f"**{bucket}**"
            lines.append(f'| {e["relative_path"]} | {e["file_type"]} | {sz} | {e["extraction_status"]} | {bucket} |')
        return '\n'.join(lines)

    def _count_stats(self, entries):
        return {
            'total': len(entries),
            'images': sum(1 for e in entries if e['file_type']=='image'),
            'texts': sum(1 for e in entries if e['file_type']=='text'),
            'sheets': sum(1 for e in entries if e['file_type']=='spreadsheet'),
            'docs': sum(1 for e in entries if e['file_type']=='document'),
            'videos': sum(1 for e in entries if e['file_type']=='video'),
            'failed': sum(1 for e in entries if e['extraction_status']=='failed'),
            'square': sum(1 for e in entries if 'square_like' in e.get('system_flags', [])),
            'collage': sum(1 for e in entries if 'possible_collage' in e.get('system_flags', [])),
        }

    def contact_sheet(self):
        manifest_path = os.path.join(self.scratch_dir, 'manifest.json')
        if not os.path.exists(manifest_path):
            print("[!] 请先运行 extract 生成 manifest.json")
            return
        if not HAS_PIL:
            print("[!] 需要 Pillow: pip install Pillow")
            return

        with open(manifest_path, 'r', encoding='utf-8') as f:
            manifest = json.load(f)

        images = [e for e in manifest if e['file_type'] == 'image']
        if not images:
            print("[!] 未找到图片")
            return

        cs_dir = os.path.join(self.scratch_dir, 'contact_sheets')
        os.makedirs(cs_dir, exist_ok=True)

        # 参数 — 放大缩略图以提高 AI 视觉识别准确率
        THUMB_W, THUMB_H = 400, 530
        COLS, ROWS = 3, 3
        PER_SHEET = COLS * ROWS  # 9
        PADDING = 10
        LABEL_H = 40
        CELL_W = THUMB_W + PADDING * 2
        CELL_H = THUMB_H + LABEL_H + PADDING * 2
        SHEET_W = CELL_W * COLS + PADDING
        SHEET_H = CELL_H * ROWS + PADDING

        from PIL import ImageDraw, ImageFont
        try:
            font = ImageFont.truetype("arial.ttf", 12)
        except Exception:
            font = ImageFont.load_default()

        # 提示颜色映射
        HINT_COLORS = {
            'ALLOW_HINT': (34, 197, 94),     # 绿
            'REVIEW_HINT': (255, 165, 0),    # 橙
            'DISCARD_HINT': (255, 0, 0),     # 红
        }

        index_lines = [
            f'# Contact Sheet 索引: {self.sku_name}', '',
            f'> 本表仅为辅助信息，图片最终分类必须由 AI 根据内容视觉判断并写入 image_selection_plan.json', '',
            '| 编号 | 文件名 | 尺寸 | 宽高比 | 路径提示 | 辅助标记 | 备注 |',
            '|:---|:---|:---|:---|:---|:---|:---|'
        ]

        # JSON 索引
        json_index = []

        # 统计
        counts = {'ALLOW_HINT': 0, 'REVIEW_HINT': 0, 'DISCARD_HINT': 0}

        sheet_idx = 0
        for batch_start in range(0, len(images), PER_SHEET):
            batch = images[batch_start:batch_start + PER_SHEET]
            sheet_idx += 1
            sheet = Image.new('RGB', (SHEET_W, SHEET_H), (255, 255, 255))
            draw = ImageDraw.Draw(sheet)

            for i, entry in enumerate(batch):
                col = i % COLS
                row = i // COLS
                x = PADDING + col * CELL_W
                y = PADDING + row * CELL_H
                global_idx = batch_start + i + 1
                tag = f'#{global_idx:03d}'

                sys_flags = entry.get('system_flags', [])
                if 'from_1x1_folder' in sys_flags or 'temp_or_unstable_path' in sys_flags:
                    status_hint = 'DISCARD_HINT'
                elif 'possible_collage' in sys_flags or 'wide_image' in sys_flags or 'square_like' in sys_flags or 'filename_prompt_like' in sys_flags:
                    status_hint = 'REVIEW_HINT'
                else:
                    status_hint = 'ALLOW_HINT'

                color = HINT_COLORS.get(status_hint, (128, 128, 128))

                # 加载缩略图
                try:
                    with Image.open(entry['absolute_path']) as img:
                        img.thumbnail((THUMB_W, THUMB_H), Image.LANCZOS)
                        ox = x + (THUMB_W - img.width) // 2 + PADDING
                        oy = y + (THUMB_H - img.height) // 2 + PADDING
                        sheet.paste(img, (ox, oy))
                except Exception:
                    draw.rectangle([x + PADDING, y + PADDING,
                                    x + THUMB_W + PADDING, y + THUMB_H + PADDING],
                                   fill=(200, 200, 200))
                    draw.text((x + PADDING + 4, y + PADDING + 4), "LOAD ERR", fill=(255, 0, 0), font=font)

                # 状态边框
                if status_hint in ('DISCARD_HINT', 'REVIEW_HINT'):
                    for offset in range(3):
                        draw.rectangle([x + offset, y + offset,
                                        x + CELL_W - offset - PADDING, y + CELL_H - offset - PADDING],
                                       outline=color)
                    draw.text((x + PADDING + 4, y + PADDING + 4), status_hint.replace('_HINT', ''), fill=color, font=font)

                # 标签：只显示编号（完整文件名保留在 index.md / index.json）
                label = tag
                draw.text((x + PADDING, y + THUMB_H + PADDING + 4), label,
                          fill=(0, 0, 0), font=font)

                # Markdown 索引条目（完整文件名保留在索引中）
                fn = os.path.basename(entry['relative_path'])
                w = entry.get('width', '?')
                h = entry.get('height', '?')
                ar = entry.get('aspect_ratio', '?')
                hints = ','.join(entry.get('path_hints', []))
                flags = ','.join(sys_flags)
                reason = entry.get('needs_review_reason', '')
                index_lines.append(f'| {tag} | `{fn}` | {w}x{h} | {ar} | {hints} | {flags} | {reason} |')

                # JSON 索引条目
                json_entry = {
                    'id': tag,
                    'relative_path': entry['relative_path'],
                    'filename': fn,
                    'width': w,
                    'height': h,
                    'aspect_ratio': ar,
                    'path_hints': entry.get('path_hints', []),
                    'system_flags': sys_flags,
                    'needs_review_reason': reason,
                    'source_root': self.target_dir,
                    'exists': os.path.exists(entry['absolute_path'])
                }
                json_index.append(json_entry)
                counts[status_hint] = counts.get(status_hint, 0) + 1

            sheet_path = os.path.join(cs_dir, f'contact_sheet_{sheet_idx:02d}.jpg')
            sheet.save(sheet_path, 'JPEG', quality=85)
            print(f"[+] Contact sheet {sheet_idx}: {len(batch)} 张图 → {sheet_path}")

        # 写入 MD 索引
        index_lines.append('')
        index_lines.append(f'**统计**: ALLOW_HINT={counts["ALLOW_HINT"]} | REVIEW_HINT={counts["REVIEW_HINT"]} | DISCARD_HINT={counts["DISCARD_HINT"]}')
        idx_path = os.path.join(cs_dir, 'contact_sheet_index.md')
        with open(idx_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(index_lines))

        # 写入 JSON 索引
        json_path = os.path.join(cs_dir, 'contact_sheet_index.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_index, f, ensure_ascii=False, indent=2)

        print(f"[+] MD 索引: {idx_path}")
        print(f"[+] JSON 索引: {json_path}")
        print(f"[+] Contact sheet 完成: {len(images)} 张图 / {sheet_idx} 张拼板")
        print(f"    ALLOW_HINT={counts['ALLOW_HINT']} | REVIEW_HINT={counts['REVIEW_HINT']} | DISCARD_HINT={counts['DISCARD_HINT']}")

    def image_candidates(self):
        # 十重安全强校验！如果校验失败，报错并强行终止 image-candidates 动作
        if not self.validate_manual_overrides(silent=False):
            print("[ERR] 错误: 由于人工干预通道校验失败，强制阻断 image-candidates 动作的生成！")
            sys.exit(1)

        overrides_path = os.path.join(self.scratch_dir, 'manual_candidate_overrides.json')
        overrides = []
        if os.path.exists(overrides_path):
            try:
                with open(overrides_path, 'r', encoding='utf-8') as f:
                    overrides = json.load(f)
            except Exception as e:
                print(f"[!] 读取人工干预通道失败: {e}")

        cs_idx_path = os.path.join(self.scratch_dir, 'contact_sheets', 'contact_sheet_index.json')
        if not os.path.exists(cs_idx_path):
            print("[!] 请先运行 contact-sheet 生成 index")
            return
        with open(cs_idx_path, 'r', encoding='utf-8') as f:
            cs_idx = json.load(f)

        manifest_path = os.path.join(self.scratch_dir, 'manifest.json')
        manifest = []
        if os.path.exists(manifest_path):
            with open(manifest_path, 'r', encoding='utf-8') as f:
                manifest = json.load(f)

        # === 通用硬排除条件（仅以下情况才从候选池中排除） ===
        HARD_EXCLUDE_FLAGS = ['temp_or_unstable_path']

        def _is_hard_excluded(item, allow_1x1_for_color=False):
            """判断是否应被硬排除。
            - temp_or_unstable_path：始终排除
            - from_1x1_folder：除非 allow_1x1_for_color=True（5.1 颜色例外）
            - possible_collage + AI 未确认：仅作 REVIEW_HINT，不硬排除
            - filename_prompt_like：不硬排除，标记 REVIEW_HINT
            """
            sys_flags = item.get('system_flags', [])
            if 'temp_or_unstable_path' in sys_flags:
                return True
            if 'from_1x1_folder' in sys_flags and not allow_1x1_for_color:
                return True
            return False

        def _needs_review(item):
            """返回该候选是否需要 AI 额外复核（REVIEW_HINT 级别）"""
            sys_flags = item.get('system_flags', [])
            review_triggers = ['possible_collage', 'wide_image', 'square_like', 'filename_prompt_like']
            return any(f in sys_flags for f in review_triggers)

        # === 提取 5.1 颜色展示候选 ===
        color_candidates = []
        for item in cs_idx:
            hints = item.get('path_hints', [])
            sys_flags = item.get('system_flags', [])
            rp = item['relative_path']
            if 'excel_images/' in rp.replace('\\', '/'):
                continue
            if '颜色' in hints:
                # 颜色候选允许 1x1 文件夹（静物/色卡例外）
                if not _is_hard_excluded(item, allow_1x1_for_color=True):
                    # 排除棚拍/外景路径提示（这些不是颜色静物图）
                    if '棚拍' not in hints and '外景' not in hints:
                        color_candidates.append(item)

        # === 提取 5.2 棚拍候选（不再硬排除 prompt 文件名）===
        studio_candidates = []
        for item in cs_idx:
            hints = item.get('path_hints', [])
            sys_flags = item.get('system_flags', [])
            if '棚拍' in hints:
                if not _is_hard_excluded(item):
                    # possible_collage 和 wide_image 不硬排除，作为 REVIEW_HINT
                    studio_candidates.append(item)

        # === 提取 5.3 外景候选（不再硬排除 prompt/midjourney 文件名）===
        outdoor_scenarios = {}
        for item in cs_idx:
            rp = item['relative_path']
            sys_flags = item.get('system_flags', [])
            hints = item.get('path_hints', [])

            if '外景' in hints:
                if not _is_hard_excluded(item):
                    # 提取子目录名称作为场景
                    parts = rp.replace('\\', '/').split('/')
                    scenario_name = '默认场景'
                    for i, p in enumerate(parts):
                        if '外景' in p and i + 1 < len(parts) - 1:
                            scenario_name = parts[i+1] + "场景"
                            if "场景场景" in scenario_name: scenario_name = scenario_name.replace("场景场景", "场景")
                            break

                    if scenario_name not in outdoor_scenarios:
                        outdoor_scenarios[scenario_name] = []
                    outdoor_scenarios[scenario_name].append(item)

        # === 提取 5.4 搭配候选 ===
        outfit_plans = []
        outfit_dirs = {}
        for item in cs_idx:
            if _is_hard_excluded(item):
                continue
            rp = item['relative_path']
            parts = rp.replace('\\', '/').split('/')
            for p in parts:
                if '搭配' in p and p != '搭配师':
                    if p not in outfit_dirs:
                        outfit_dirs[p] = []
                    outfit_dirs[p].append(item)

        # === 物理级人工 overrides 内存合流 (Step 4) ===
        for entry in overrides:
            pid = entry['id']
            orig_path = entry['original_path']
            section = entry['target_section']
            created_by = entry.get('created_by', 'manual')
            override_id = entry.get('manual_override_id', f"#manual_{pid}")

            clean_orig_path = orig_path.replace('\\', '/')
            abs_fpath = os.path.abspath(os.path.join(self.target_dir, clean_orig_path))

            img_meta = {"width": 0, "height": 0, "channels": 3}
            try:
                # 尝试从 asset_state.json 获取分析结果，若无，则直接分析
                state_path = os.path.join(self.scratch_dir, 'asset_state.json')
                if os.path.exists(state_path):
                    with open(state_path, 'r', encoding='utf-8') as sf:
                        state_data = json.load(sf)
                        if clean_orig_path in state_data:
                            old_e = state_data[clean_orig_path]
                            img_meta["width"] = old_e.get("width", 0)
                            img_meta["height"] = old_e.get("height", 0)
                            img_meta["channels"] = old_e.get("channels", 3)
                if img_meta["width"] == 0 and os.path.exists(abs_fpath):
                    img_meta.update(self._analyze_image(abs_fpath, clean_orig_path))
            except Exception as ex:
                print(f"[!] 自动解析 override 图像尺寸失败: {clean_orig_path}, 错误: {ex}")

            mock_item = {
                "id": pid,
                "relative_path": clean_orig_path,
                "width": img_meta.get("width", 0),
                "height": img_meta.get("height", 0),
                "channels": img_meta.get("channels", 3),
                "source": "manual_override",
                "manual_override_id": override_id,
                "requires_original_review": True,
                "path_hints": ["人工干预"],
                "system_flags": ["manual_override"]
            }

            # 内存安全合流分支分配
            if section == "5.1 颜色展示":
                color_candidates.append(mock_item)
            elif section == "5.2 棚拍展示":
                studio_candidates.append(mock_item)
            elif section == "5.3 外景展示":
                scenario_name = entry.get("target_scenario", "默认场景")
                if "场景" not in scenario_name:
                    scenario_name += "场景"
                if "场景场景" in scenario_name:
                    scenario_name = scenario_name.replace("场景场景", "场景")
                if scenario_name not in outdoor_scenarios:
                    outdoor_scenarios[scenario_name] = []
                outdoor_scenarios[scenario_name].append(mock_item)
            elif section == "5.4 搭配方案展示":
                outfit_name = entry.get("target_outfit_name")
                if not outfit_name:
                    parts = clean_orig_path.split('/')
                    outfit_name = "默认搭配"
                    for p in parts:
                        if "搭配" in p and p != "搭配师":
                            outfit_name = p
                            break
                if outfit_name not in outfit_dirs:
                    outfit_dirs[outfit_name] = []
                outfit_dirs[outfit_name].append(mock_item)
            elif section == "8.1 面料与品牌资质":
                fabric_candidates.append(mock_item)

        for oname, items in outfit_dirs.items():
            outfit_plans.append({
                "name": oname,
                "candidate_count": len(items),
                "candidate_ids": [i['id'] for i in items],
                "note": "AI 需打开原图判断每个单品的 observed_object、item_type 和 display_name"
            })

        # === 提取 8.1 面料与品牌资质 ===
        FABRIC_PATH_KEYWORDS = ('面料', '质感', '细节', '工艺', 'swatch', 'fabric', '纹理', '编织', '检测报告')
        fabric_candidates = []
        for item in cs_idx:
            hints = item.get('path_hints', [])
            rp = item['relative_path']
            rp_lower = rp.lower()
            sys_flags = item.get('system_flags', [])
            # 匹配条件：path_hints 或 relative_path 包含面料相关关键词，或是 Excel 自动提取的图片
            hint_match = any(h in hints for h in ('面料', '质感', '细节', '工艺'))
            path_match = any(kw in rp_lower for kw in FABRIC_PATH_KEYWORDS) or 'excel_images/' in rp_lower
            # '基础物料' 单独匹配，排除搭配路径里的物料
            base_material_match = '基础物料' in rp_lower and '搭配' not in rp_lower
            is_pdf_thumb = self.stable_assets_rel in rp.replace('\\', '/')
            if hint_match or path_match or base_material_match or is_pdf_thumb:
                if 'temp_or_unstable_path' not in sys_flags and '棚拍' not in hints and '外景' not in hints and '搭配' not in hints:
                    fabric_candidates.append(item)

        # === 组装候选池 JSON ===
        candidates = {
            "5.1 颜色展示": {
                "candidate_count": len(color_candidates),
                "candidate_ids": [i['id'] for i in color_candidates],
                "note": "候选池，AI 需打开原图确认为产品静物图/色卡图/正反面图/多色对比图，禁止模特图。所有确认为有效颜色图的候选都应入选（coverage_mode=all_valid_candidates），不得用 Excel/文本数据的颜色字段反向过滤。非颜色图（如检测报告、面料样卡）需写入 excluded_candidates 说明排除原因"
            },
            "5.2 棚拍展示": {
                "candidate_count": len(studio_candidates),
                "candidate_ids": [i['id'] for i in studio_candidates],
                "note": "候选池，AI 需打开原图逐一复核后决定最终选用"
            },
            "5.3 外景展示": {},
            "5.4 搭配方案展示": outfit_plans,
            "8.1 面料与品牌资质": {
                "candidate_count": len(fabric_candidates),
                "candidate_ids": [i['id'] for i in fabric_candidates],
                "note": "候选池，AI 需打开原图确认是面料/工艺/成分标签/报告，禁止普通模特图"
            }
        }
        for sname, items in outdoor_scenarios.items():
            candidates["5.3 外景展示"][sname] = {
                "candidate_count": len(items),
                "candidate_ids": [i['id'] for i in items],
                "note": f"候选池，AI 需打开原图逐一复核 {sname} 场景图"
            }

        out_json = os.path.join(self.scratch_dir, 'image_candidates.json')
        with open(out_json, 'w', encoding='utf-8') as f:
            json.dump(candidates, f, ensure_ascii=False, indent=2)

        # === 生成 MD 报告 ===
        md_lines = [
            '# 图像候选池', '',
            '> 💡 本文件由 `image-candidates` 自动生成，反映了当前目录中的客观图片情况。',
            '> `candidate_count` 是候选池中的图片数量，不是最终手册必须展示的数量。',
            '> AI 必须打开候选原图逐一复核，最终选用数量和理由写入 `image_selection_plan.json`。', '',
            '> ⚠️ `image_candidates.json` 是引擎输出，禁止手动编辑修复候选池。如有漏图必须修通用逻辑后重跑。', '',
            f'## 5.1 颜色展示',
            f'- **候选数量 (candidate_count)**: {len(color_candidates)}',
            f'- **候选图 ID**: {", ".join([i["id"] for i in color_candidates]) if color_candidates else "无"}', '',
            f'## 5.2 棚拍展示',
            f'- **候选数量 (candidate_count)**: {len(studio_candidates)}',
            f'- **候选图 ID**: {", ".join([i["id"] for i in studio_candidates]) if studio_candidates else "无"}', '',
            '## 5.3 外景展示'
        ]
        if not outdoor_scenarios:
            md_lines.append('- 无外景场景')
        for sname, items in outdoor_scenarios.items():
            md_lines.extend([
                f'### {sname}',
                f'- **候选数量 (candidate_count)**: {len(items)}',
                f'- **候选图 ID**: {", ".join([i["id"] for i in items])}', ''
            ])

        md_lines.append('## 5.4 搭配方案展示')
        if not outfit_plans:
            md_lines.append('- 无搭配方案')
        for plan in outfit_plans:
            md_lines.extend([
                f'### {plan["name"]}',
                f'- **候选数量**: {plan["candidate_count"]}',
                f'- **候选图 ID**: {", ".join(plan["candidate_ids"])}',
                f'- AI 需打开原图判断 observed_object、item_type 和 display_name', ''
            ])

        md_lines.extend([
            '## 8.1 面料与品牌资质',
            f'- **候选数量 (candidate_count)**: {len(fabric_candidates)}',
            f'- **候选图 ID**: {", ".join([i["id"] for i in fabric_candidates]) if fabric_candidates else "无"}', '',
            '## 🛡️ 人工干预与审计追踪表 (Manual Overrides Audit Chain)', ''
        ])

        if overrides:
            md_lines.extend([
                '| 候选项 ID | 物理路径 | 目标小节 | 干预 ID | 来源渠道 | 强制原图复核 |',
                '| :--- | :--- | :--- | :--- | :--- | :--- |'
            ])
            for entry in overrides:
                pid = entry['id']
                orig_path = entry['original_path']
                section = entry['target_section']
                created_by = entry.get('created_by', 'manual')
                override_id = entry.get('manual_override_id', f"#manual_{pid}")
                md_lines.append(f'| {pid} | {orig_path} | {section} | {override_id} | {created_by} | **TRUE** |')
            md_lines.append('')
        else:
            md_lines.extend([
                '> 💡 当前未引入任何人工干预配置，全部候选来自系统自动物理扫描。', ''
            ])

        out_md = os.path.join(self.scratch_dir, 'image_candidates.md')
        with open(out_md, 'w', encoding='utf-8') as f:
            f.write('\n'.join(md_lines))

        print(f"[+] 图像候选分析已生成:")
        print(f"    - {out_json}")
        print(f"    - {out_md}")

    def load_launch_scope_from_excel(self) -> dict:
        """
        v2.3.9.1 launch-scope-filter hotfix:
        从上新资料包 Excel 表格中加载允许的颜色与尺码白名单。
        """
        launch_scope = {
            "detected": False,
            "source_file": "None",
            "source_type": "None",
            "allowed_colors": [],
            "allowed_color_aliases": {},
            "allowed_sizes": [],
            "excluded_colors": [],
            "excluded_sizes": [],
            "policy": {
                "color_filter_mode": "excel_whitelist",
                "size_filter_mode": "excel_whitelist",
                "if_excel_missing": "skip_filter",
                "if_excel_parse_failed": "warn_and_skip"
            }
        }

        # 1. 扫描根目录下所有的 .xlsx 表格文件，匹配含有 "上新"、"资料包" 或 "属性" 的 Excel 文件
        xlsx_file = None
        for filename in os.listdir(self.target_dir):
            if filename.endswith(".xlsx") and not filename.startswith("~$"):
                if any(x in filename for x in ("上新", "资料包", "属性", "基本", "商品")):
                    xlsx_file = os.path.join(self.target_dir, filename)
                    break

        # 兜底：如果没找到符合名字的，只要是第一个非临时的 .xlsx 都可以
        if not xlsx_file:
            for filename in os.listdir(self.target_dir):
                if filename.endswith(".xlsx") and not filename.startswith("~$"):
                    xlsx_file = os.path.join(self.target_dir, filename)
                    break

        # 如果有老式 .xls 提示 warning
        for filename in os.listdir(self.target_dir):
            if filename.endswith(".xls"):
                safe_print(f"[WARN] 发现老式表格文件 '{filename}'。当前版本暂不解析 .xls，请另存为 .xlsx 后重试。")

        if not xlsx_file:
            safe_print("[*] 商品目录下未检索到产品上新资料包 Excel。优雅跳过白名单过滤（Skip Filter）。")
            return launch_scope

        safe_print(f"[*] 检索到上新资料包: {os.path.basename(xlsx_file)}")

        # 2. 尝试载入并解析属性
        if not HAS_OPENPYXL:
            safe_print("[WARN] 未安装 openpyxl，无法解析 Excel 白名单，跳过过滤。")
            return launch_scope

        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
            ws = wb.worksheets[0] # 读取第一个 sheet

            raw_colors_str = ""
            raw_sizes_str = ""

            # 使用更宽泛的别名匹配字段名和属性行
            color_headers = ["颜色", "颜色名称", "色名", "色号", "color", "Color"]
            size_headers = ["尺码", "尺码范围", "规格", "尺码规格", "size", "Size"]

            # 扫描 Sheet 1 的每一行，键值对行级横向匹配
            for row in ws.iter_rows(values_only=True):
                if not row or not any(row):
                    continue
                row_clean = [str(c).strip() if c is not None else '' for c in row]

                key_col = row_clean[0]
                if any(ch in key_col for ch in color_headers) and len(row_clean) > 1:
                    raw_colors_str = row_clean[1]
                if any(sh in key_col for sh in size_headers) and len(row_clean) > 1:
                    raw_sizes_str = row_clean[1]

            wb.close()

            # 3. 解析允许展示的颜色与尺码
            allowed_colors = []
            if raw_colors_str:
                parts = re.split(r'[/,，;；、\s]+', raw_colors_str)
                allowed_colors = [p.strip() for p in parts if p.strip()]

            allowed_sizes = []
            if raw_sizes_str:
                parts = re.split(r'[/,，;；、\s]+', raw_sizes_str)
                allowed_sizes = [p.strip() for p in parts if p.strip()]

            if not allowed_colors and not allowed_sizes:
                safe_print("[WARN] 上新资料包中未提取到有效的颜色或尺码范围属性，跳过过滤。")
                return launch_scope

            # 4. 读取外置的 attribute_aliases.json 别名配置
            color_aliases = {}
            size_aliases = {}
            aliases_path = os.path.join(self.skill_dir, "references", "attribute_aliases.json")
            if os.path.exists(aliases_path):
                try:
                    with open(aliases_path, 'r', encoding='utf-8') as af:
                        aliases_data = json.load(af)
                        color_aliases = aliases_data.get("color_aliases", {})
                        size_aliases = aliases_data.get("size_aliases", {})
                except Exception as ae:
                    safe_print(f"[WARN] 加载别名配置错误: {ae}")

            # 5. 生成允许的颜色别名映射
            allowed_color_aliases = {}
            all_allowed_colors_and_aliases = set(allowed_colors)
            for c in allowed_colors:
                # 检查同义词表，如果当前颜色存在同义词映射
                if c in color_aliases:
                    allowed_color_aliases[c] = color_aliases[c]
                    for alias in color_aliases[c]:
                        all_allowed_colors_and_aliases.add(alias)
                # 反向查找：如果允许的颜色是别名，比如允许 "咖色"
                for main_color, alias_list in color_aliases.items():
                    if c in alias_list:
                        all_allowed_colors_and_aliases.add(main_color)
                        for other in alias_list:
                            all_allowed_colors_and_aliases.add(other)

            # 6. 计算被淘汰（禁用）的颜色与尺码
            # 定义全局可能出现的候选颜色与标准尺码
            common_colors = ["黑色", "白色", "红色", "黄色", "蓝色", "绿色", "粉色", "杏色", "米色", "燕麦色", "驼色", "灰色", "紫色", "咖色", "咖啡色", "棕色", "藏青", "深灰"]
            common_sizes = ["XS", "S", "M", "L", "XL", "XXL", "XXXL", "均码"]

            excluded_colors = [c for c in common_colors if c not in all_allowed_colors_and_aliases]
            excluded_sizes = [s for s in common_sizes if s not in allowed_sizes]

            # 组装 launch_scope json
            launch_scope.update({
                "detected": True,
                "source_file": os.path.basename(xlsx_file),
                "source_type": "launch_sheet",
                "allowed_colors": allowed_colors,
                "allowed_color_aliases": allowed_color_aliases,
                "allowed_sizes": allowed_sizes,
                "excluded_colors": excluded_colors,
                "excluded_sizes": excluded_sizes
            })

            # 保存 JSON
            ls_path = os.path.join(self.scratch_dir, "launch_scope.json")
            with open(ls_path, 'w', encoding='utf-8') as lsf:
                json.dump(launch_scope, lsf, ensure_ascii=False, indent=2)

            # 7. 生成人眼可读的过滤报告 launch_scope_filter_report.md
            reports_dir = os.path.join(self.scratch_dir, "reports")
            os.makedirs(reports_dir, exist_ok=True)
            report_path = os.path.join(reports_dir, "launch_scope_filter_report.md")

            rep_lines = [
                "# 📊 上新资料包颜色尺码过滤报告",
                "",
                "## 🖥️ 1. 识别到的上新配置",
                f"* **资料包文件:** `{os.path.basename(xlsx_file)}`",
                f"* **检测状态:** `ON (启用白名单对齐过滤) 🟢`",
                "",
                "## 🎨 2. 允许展示的可售颜色",
                "最终手册将**只允许**在结构化区域展示以下颜色及映射别名：",
                f"* **标准上新色:** {', '.join([f'`{x}`' for x in allowed_colors])}",
                f"* **支持的同义别名:** {', '.join([f'`{a}`' for a in all_allowed_colors_and_aliases - set(allowed_colors)]) if (all_allowed_colors_and_aliases - set(allowed_colors)) else '无'}",
                "",
                "## 📏 3. 允许展示的生产尺码",
                "最终手册将**只允许**在结构化区域展示以下下单尺码：",
                f"* **标准下单尺码:** {', '.join([f'`{x}`' for x in allowed_sizes])}",
                "",
                "## 🚫 4. 被成功过滤的属性大盘",
                "| 属性类型 | 被过滤清除项 | 过滤原因 |",
                "|---|---|---|",
            ]
            for ec in excluded_colors:
                rep_lines.append(f"| 颜色 | `{ec}` | 上新资料包 Excel 中未包含，物理安全隔离清除 |")
            for es in excluded_sizes:
                rep_lines.append(f"| 尺码 | `{es}` | 上新资料包 Excel 中未包含，尺码表物理整行删除 |")

            rep_lines.append("\n## 🚦 5. 过滤结论\n最终交付的 Markdown 手册和 HTML 编译视图已严格对齐产品下单计划，自动剔除了未下单的废弃属性！")

            with open(report_path, 'w', encoding='utf-8') as rf:
                rf.write("\n".join(rep_lines))

            safe_print(f"[SUCCESS] v2.3.9.1 上新白名单生成成功！")
            safe_print(f"  - 允许颜色: {allowed_colors}")
            safe_print(f"  - 允许尺码: {allowed_sizes}")
            safe_print(f"  - 过滤报告已输出: {report_path}")

        except Exception as e:
            safe_print(f"[WARN] 解析上新 Excel 白名单发生错误 (优雅降级): {e}")

        return launch_scope

    def _extract_images_from_all_xlsx(self):
        """扫描 target_dir 内的 xlsx 文件，提取其中的常规图片与 cellImages(DISPIMG) 到 excel_images/ 文件夹"""
        excel_images_dir = os.path.join(self.target_dir, 'excel_images')

        # 每次提取前清空 excel_images 目录，以防历史脏数据残留
        if os.path.exists(excel_images_dir):
            try:
                import shutil
                shutil.rmtree(excel_images_dir)
            except Exception as e:
                safe_print(f"[WARN] 清理 excel_images 目录失败: {e}")

        os.makedirs(excel_images_dir, exist_ok=True)

        skip_dirs = {'.scratch', '自动生成素材', 'scratch', 'parts', '.git', '__pycache__', 'excel_images'}

        for root, dirs, files in os.walk(self.target_dir):
            dirs[:] = [d for d in dirs if d not in skip_dirs]
            for fname in files:
                if fname.lower().endswith('.xlsx') and not fname.startswith('~$'):
                    fpath = os.path.join(root, fname)
                    try:
                        self._extract_images_from_single_xlsx(fpath, excel_images_dir)
                    except Exception as ex:
                        safe_print(f"[WARN] 从 Excel {fname} 提取图片失败: {ex}")

    def _extract_images_from_single_xlsx(self, xlsx_path, out_dir):
        """解析单张 xlsx 提取图片，支持常规嵌入图片与 cellImages (DISPIMG) 映射"""
        import zipfile
        import re

        if not zipfile.is_zipfile(xlsx_path):
            return

        with zipfile.ZipFile(xlsx_path, 'r') as z:
            namelist = z.namelist()

            # 1) 解析 cellimages.xml.rels 建立 rId -> target_path 映射
            rels_map = {}
            rels_path = 'xl/_rels/cellimages.xml.rels'
            if rels_path in namelist:
                try:
                    rels_data = z.read(rels_path).decode('utf-8', errors='ignore')
                    # 匹配 <Relationship Id="rId1" Type="..." Target="media/image1.png"/>
                    matches = re.findall(r'Id=["\']([^"\']+)["\'][^>]*Target=["\']([^"\']+)["\']', rels_data)
                    for rid, target in matches:
                        if not target.startswith('xl/'):
                            if target.startswith('media/'):
                                target = 'xl/' + target
                            else:
                                target = 'xl/media/' + target
                        rels_map[rid] = target
                except Exception as e:
                    safe_print(f"[WARN] 读取 xl/_rels/cellimages.xml.rels 发生错误: {e}")

            # 2) 解析 cellimages.xml 建立 rId -> wps_name (ID_XXX) 映射
            wps_name_map = {}
            cellimages_path = 'xl/cellimages.xml'
            if cellimages_path in namelist:
                try:
                    cellimages_data = z.read(cellimages_path).decode('utf-8', errors='ignore')
                    blocks = re.findall(r'<etc:cellImage\b[^>]*>(.*?)</etc:cellImage>', cellimages_data, re.S)
                    for blk in blocks:
                        m_name = re.search(r'name=["\'](ID_[A-Za-z0-9]+)["\']', blk)
                        m_embed = re.search(r'embed=["\']([^"\']+)["\']', blk)
                        if m_name and m_embed:
                            wps_name_map[m_embed.group(1)] = m_name.group(1)
                except Exception as e:
                    safe_print(f"[WARN] 读取 xl/cellimages.xml 发生错误: {e}")

            # 3) 遍历 xl/media/ 提取图片
            for name in namelist:
                if name.startswith('xl/media/') and not name.endswith('/'):
                    ext = os.path.splitext(name)[1].lower()
                    if ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp'):
                        # 匹配 wps_id
                        matched_rids = [rid for rid, target in rels_map.items() if target == name]
                        wps_id = None
                        if matched_rids:
                            for rid in matched_rids:
                                if rid in wps_name_map:
                                    wps_id = wps_name_map[rid]
                                    break

                        # 确定输出文件名
                        if wps_id:
                            out_name = f"excel_img_{wps_id}{ext}"
                        else:
                            base = os.path.basename(name)
                            out_name = f"excel_img_{base}"

                        out_path = os.path.join(out_dir, out_name)
                        try:
                            with open(out_path, 'wb') as out_f:
                                out_f.write(z.read(name))
                            safe_print(f"[*] 成功提取 Excel 图片: {name} -> {out_name}")
                        except Exception as e:
                            safe_print(f"[WARN] 写入提取的图片 {out_name} 失败: {e}")
