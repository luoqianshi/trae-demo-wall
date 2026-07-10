"""报告生成器 - 导出 HTML / Markdown / Excel 格式报告"""
import os
import datetime
from typing import Dict, List

class ReportGenerator:
    def __init__(self, export_dir: str):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)

    def export_html(self, session: Dict, report: Dict) -> str:
        """导出完整 HTML 报告"""
        summary = report['summary']
        total = summary['pass'] + summary['pending'] + summary['fail']
        pass_pct = round(summary['pass'] / total * 100) if total else 0

        items_html = ''
        for item in report['items']:
            badge = {'符合': 'badge-pass', '待确认': 'badge-pending', '不符合': 'badge-fail'}
            cls = badge.get(item['result'], 'badge-pending')
            risk = '<span class="risk">🔴 高风险</span>' if item['severity'] == 'high' and item['result'] == '不符合' else ''
            sug = f'<div class="sug"><strong>改进建议:</strong> {item["suggestion"]}</div>' if item.get('suggestion') else ''
            ai = f'<div class="ai-box"><strong>AI 分析:</strong> {item["ai_analysis"]}</div>' if item.get('ai_analysis') else ''
            items_html += f'''
            <div class="item">
                <div class="item-header"><span class="id">{item["id"]}</span><span class="title">{item["check_item"]}</span>{risk}<span class="{cls}">{item["result"]}</span></div>
                <div class="item-body"><div class="row"><span class="label">证据</span><span>{item["evidence"]}</span></div><div class="row"><span class="label">标准</span><span>{item["standard_ref"]}</span></div>{sug}{ai}</div>
            </div>'''

        html = f'''<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="UTF-8"><title>审核报告 - {report['project_name']}</title>
<style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif; background: #f0f4f8; color: #1a202c; padding: 40px 20px; }}
    .container {{ max-width: 900px; margin: 0 auto; }}
    .header {{ text-align: center; margin-bottom: 30px; }}
    .header h1 {{ font-size: 24px; color: #1a365d; }}
    .header p {{ color: #718096; font-size: 14px; margin-top: 4px; }}
    .summary {{ display: flex; gap: 12px; justify-content: center; margin-bottom: 30px; }}
    .summary-card {{ text-align: center; padding: 16px 24px; border-radius: 10px; border: 1px solid #e2e8f0; min-width: 100px; }}
    .summary-card .num {{ font-size: 28px; font-weight: 700; }}
    .summary-card .label {{ font-size: 13px; color: #718096; }}
    .sc-pass .num {{ color: #38a169; }} .sc-pending .num {{ color: #d69e2e; }} .sc-fail .num {{ color: #e53e3e; }}
    .item {{ border: 1px solid #e2e8f0; border-radius: 8px; margin-bottom: 10px; overflow: hidden; }}
    .item-header {{ display: flex; align-items: center; gap: 10px; padding: 12px 16px; background: white; }}
    .item-header .id {{ font-size: 12px; color: #a0aec0; font-weight: 600; min-width: 100px; }}
    .item-header .title {{ flex: 1; font-size: 14px; color: #2d3748; }}
    .item-body {{ padding: 12px 16px; border-top: 1px solid #e2e8f0; font-size: 13px; }}
    .item-body .row {{ margin-bottom: 6px; display: flex; gap: 8px; }}
    .item-body .label {{ min-width: 50px; color: #a0aec0; font-weight: 500; }}
    .badge-pass {{ background: #c6f6d5; color: #22543d; padding: 2px 10px; border-radius: 12px; font-size: 12px; font-weight: 600; }}
    .badge-pending {{ background: #fefcbf; color: #744210; padding: 2px 10px; border-radius: 12px; font-size: 12px; font-weight: 600; }}
    .badge-fail {{ background: #fed7d7; color: #9b2c2c; padding: 2px 10px; border-radius: 12px; font-size: 12px; font-weight: 600; }}
    .risk {{ color: #c53030; font-size: 12px; font-weight: 600; }}
    .ai-box {{ background: #fffbeb; border: 1px solid #f6e05e; border-radius: 6px; padding: 8px 12px; margin-top: 8px; font-size: 13px; }}
    .sug {{ background: #ebf8ff; border: 1px solid #90c4f9; border-radius: 6px; padding: 8px 12px; margin-top: 8px; font-size: 13px; }}
    .footer {{ text-align: center; margin-top: 30px; color: #a0aec0; font-size: 12px; }}
</style></head><body>
<div class="container">
    <div class="header"><h1>📋 审核报告</h1><p>{report['project_name']} | {report['process_area']} | {report['target_level']} | {report['review_time']}</p></div>
    <div class="summary">
        <div class="summary-card sc-pass"><div class="num">{summary['pass']}</div><div class="label">✅ 符合</div></div>
        <div class="summary-card sc-pending"><div class="num">{summary['pending']}</div><div class="label">⚠️ 待确认</div></div>
        <div class="summary-card sc-fail"><div class="num">{summary['fail']}</div><div class="label">❌ 不符合</div></div>
    </div>
    <div>{items_html}</div>
    <div class="footer"><p>生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</p><p>⚠️ AI 生成结果仅供参考，需专业人员审核确认</p></div>
</div></body></html>'''

        filepath = os.path.join(self.export_dir, f'{session["session_id"]}_report.html')
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html)
        return filepath

    def export_markdown(self, session: Dict, report: Dict) -> str:
        """导出 Markdown 报告"""
        summary = report['summary']
        lines = [
            f'# 审核报告: {report["project_name"]}',
            '',
            f'- **过程域**: {report["process_area"]}',
            f'- **目标等级**: {report["target_level"]}',
            f'- **审核时间**: {report["review_time"]}',
            '',
            '## 统计汇总',
            '',
            f'| 结果 | 数量 |',
            f'|------|------|',
            f'| ✅ 符合 | {summary["pass"]} |',
            f'| ⚠️ 待确认 | {summary["pending"]} |',
            f'| ❌ 不符合 | {summary["fail"]} |',
            '',
            '## 逐项结果',
            '',
        ]
        for item in report['items']:
            lines.append(f'### {item["id"]}: {item["check_item"]}')
            lines.append(f'- **结果**: {item["result"]}')
            lines.append(f'- **证据**: {item["evidence"]}')
            lines.append(f'- **标准**: {item["standard_ref"]}')
            if item.get('suggestion'):
                lines.append(f'- **改进建议**: {item["suggestion"]}')
            if item.get('ai_analysis'):
                lines.append(f'- **AI 分析**: {item["ai_analysis"]}')
            lines.append('')

        md = '\n'.join(lines)
        filepath = os.path.join(self.export_dir, f'{session["session_id"]}_report.md')
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(md)
        return filepath

    def export_excel(self, session: Dict, report: Dict) -> str:
        """导出 Excel 报告"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '审核报告'

        # 标题
        ws.merge_cells('A1:G1')
        ws['A1'] = f'审核报告: {report["project_name"]}'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'过程域: {report["process_area"]} | 目标等级: {report["target_level"]} | 时间: {report["review_time"]}'

        # 表头
        headers = ['检查项ID', '检查内容', '标准条款', '审核结果', '证据', '改进建议', 'AI分析']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2B6CB0', end_color='2B6CB0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # 数据
        for i, item in enumerate(report['items'], 5):
            ws.cell(row=i, column=1, value=item['id'])
            ws.cell(row=i, column=2, value=item['check_item'])
            ws.cell(row=i, column=3, value=item['standard_ref'])
            ws.cell(row=i, column=4, value=item['result'])
            ws.cell(row=i, column=5, value=item['evidence'])
            ws.cell(row=i, column=6, value=item.get('suggestion', ''))
            ws.cell(row=i, column=7, value=item.get('ai_analysis', ''))

        # 列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 40

        filepath = os.path.join(self.export_dir, f'{session["session_id"]}_report.xlsx')
        wb.save(filepath)
        return filepath